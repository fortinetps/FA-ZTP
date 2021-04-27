import openpyxl
import yaml
import json
import argparse
import os
from print_schema import print_schema


def convert_worksheet_to_dict(workbook_path=None, worksheet=None):
    wb = openpyxl.load_workbook(workbook_path)
    sheet = wb[worksheet]

    # get the column names into a list so we can reference them when gathering rows
    column_names = []
    for column in sheet.iter_cols(1, sheet.max_column):
        column_names.append(column[0].value)

    # init the rows lists for each column in the dictionary
    # we'll end up with a dictionary containing lists of all row values, a column view
    # we then need to change that to a row view after
    row_data = {}
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if len(row) != len(column_names):
            print("Row length didn't match column length! Usually means table isn't formatted correctly. Exiting!")
            exit(-1)
        if i == 0:
            for column_index, column_value in enumerate(row):
                row_data[row[column_index]] = []
        else:
            for column_index, column_value in enumerate(row):
                row_data[column_names[column_index]].append(column_value)

    row_count = int(i)
    # print(row_count)

    # now convert to a list of dictionaries, rather than a dictionary of lists.
    # moving from column based to row based view
    # first make the template dict

    ROW_TEMPLATE = {}
    for column in column_names:
        ROW_TEMPLATE[column] = None

    final_row_list = []
    row_index = 0
    while row_index < row_count:
        append_dict = dict(ROW_TEMPLATE)
        for column in column_names:
            if row_data[column][row_index]:
                try:
                    if '\n' in str(row_data[column][row_index]):
                        row_value = list(
                            filter(None, row_data[column][row_index].split('\n')))
                    else:
                        row_value = row_data[column][row_index]
                except TypeError as e:
                    str(e)
                    raise
            else:
                row_value = row_data[column][row_index]
            append_dict[column] = row_value
        empty_row = True
        for k, v in append_dict.items():
            if v:
                empty_row = False

        if not empty_row:
            final_row_list.append(append_dict)
        row_index += 1

    return final_row_list


def convert_whole_workbook_to_dict(workbook_path=None):
    wb = openpyxl.load_workbook(workbook_path)
    workbook_dict = dict()
    sheet_names = wb.sheetnames
    for sheet in sheet_names:
        try:
            print("Reading: " + sheet.replace(" ", "_"))
            worksheet_dict = convert_worksheet_to_dict(worksheet=sheet,
                                                       workbook_path=workbook_path)
            workbook_dict[sheet.replace(" ", "_")] = worksheet_dict
        except Exception as e:
            str(e)
            raise
    return workbook_dict


def format_excel_columns(input_dict=None):
    ret_dict = None
    if input_dict:
        ret_dict = {}
        for k, v in input_dict.items():
            if k:
                try:
                    ret_dict[k.replace(' ', '_').lower()] = v
                except Exception as e:
                    print(e)
                    exit(-1)
    return ret_dict


def generate_fortigate_var_files(sheets=None):
    ztp_fortigates = []
    print(sheets)
    for fgt in sheets['FortiGates']:
        fgt_dict = {
            'fgt_name': fgt['FortiGate Name'],
            'fgt_sn': None,
            'fgt_mgmt_port': fgt['MGMT Port'],
            'fgt_mgmt_ip': fgt['MGMT IP'],
            'fgt_mgmt_vrf': fgt['MGMT VRF'],
            'fgt_wan1_port': fgt['WAN1 Port'],
            'fgt_wan1_ip': fgt['WAN1 IP'],
            'fgt_wan1_dl_speed_kbps': fgt['WAN1 DL kbps'],
            'fgt_wan1_ul_speed_kbps': fgt['WAN1 UL kbps'],
            'fgt_wan2_port': fgt['WAN2 Port'],
            'fgt_wan2_ip': fgt['WAN2 IP'],
            'fgt_wan2_dl_speed_kbps': fgt['WAN2 DL kbps'],
            'fgt_wan2_ul_speed_kbps': fgt['WAN2 UL kbps'],
            'fgt_wan3_port': fgt['WAN3 Port'],
            'fgt_wan3_ip': fgt['WAN3 IP'],
            'fgt_wan3_dl_speed_kbps': fgt['WAN3 DL kbps'],
            'fgt_wan3_ul_speed_kbps': fgt['WAN3 UL kbps'],
            'fmgr_target_adom': fgt['FMGR ADOM'],
            'fgt_vdom': fgt['FGT VDOM'],
            'fortilink_stack_ip': fgt['FortiLink Stack IP'],
            'fmgr_target_device_group': fgt['FMGR Device Group'],
            'fgt_local_bgp_as': None,
            'fap_profiles': [],
        }
        # 'fgt_fortiswitch_trunk_interfaces': fgt['FortiSwitch Trunk Interfaces']
        if fgt['FortiLink Trunk Interfaces']:
            if "," in fgt['FortiLink Trunk Interfaces']:
                fgt_dict['fgt_fortilink_trunk_interfaces'] = fgt['FortiLink Trunk Interfaces'].split(',')
            elif "\n" in fgt['FortiLink Trunk Interfaces']:
                fgt_dict['fgt_fortilink_trunk_interfaces'] = fgt['FortiLink Trunk Interfaces'].split('\n')
            else:
                fgt_dict['fgt_fortilink_trunk_interfaces'] = [fgt['FortiLink Trunk Interfaces']]
        else:
            fgt_dict['fgt_fortilink_trunk_interfaces'] = []

        # get the fortigate location name by device name
        for device_location in sheets['DeviceLocations']:
            if fgt_dict['fgt_name'] == device_location['Device Name']:
                fgt_dict['fgt_location_name'] = device_location['Location Name']
                fgt_dict['fgt_device_model'] = device_location['Device Model']
                fgt_dict['fgt_sn'] = device_location['Device Serial Number']

        # fill in the fortigate location data
        for location in sheets['Locations']:
            if fgt_dict['fgt_location_name'] == location['Location Name']:
                fgt_dict['fgt_street_address'] = location['Full Street Address']
                fgt_dict['fgt_latitude'] = location['Latitude']
                fgt_dict['fgt_longitude'] = location['Longitude']

        # get the template assignments if any
        fgt_dict['templates'] = []
        for fgt_template_set in sheets['Template_Assignments']:
            if fgt_dict['fgt_name'] == fgt_template_set['FortiGate Name']:
                fgt_dict['templates'].append(format_excel_columns(fgt_template_set))
        if len(fgt_dict['templates']) > 1:
            raise Exception("There is more than one row of template assignments for a FortiGate. "
                            "Audit the Template Assignments input worksheet")

        # fill in the static routes for fortigate
        fgt_dict['fgt_static_routes'] = []
        for route in sheets['Static_Routes']:
            if fgt_dict['fgt_name'] == route['FortiGate Name']:
                fgt_dict['fgt_static_routes'].append(format_excel_columns(route))

        # get the networks assigned to this fortigate
        fgt_dict['fgt_lan_vlan_map'] = []
        for network in sheets['Networks']:
            if fgt_dict['fgt_name'] == network['FortiGate Name']:
                formatted_network = format_excel_columns(network)
                if formatted_network['allow_access']:
                    if "," in formatted_network['allow_access']:
                        allow_access = formatted_network['allow_access'].split(',')
                    elif "\n" in formatted_network['allow_access']:
                        allow_access = formatted_network['allow_access'].split('\n')
                    else:
                        allow_access = [formatted_network['allow_access']]
                    formatted_network['allow_access'] = allow_access
                    fgt_dict['fgt_lan_vlan_map'].append(formatted_network)

        # create the zones
        fgt_dict['zones'] = []
        distinct_zone_list = []
        for network in fgt_dict['fgt_lan_vlan_map']:
            if network['zone'] not in distinct_zone_list and network['zone']:
                distinct_zone_list.append(network['zone'])
        for zone in distinct_zone_list:
            net_dict = {'name': zone, 'members': []}
            for network in fgt_dict['fgt_lan_vlan_map']:
                if network['zone'] == zone:
                    net_dict['members'].append(network['network_alias'])
            fgt_dict['zones'].append(net_dict)

        # get any switches downstream from this fortigate
        fgt_dict['fortiswitches'] = []
        for device in sheets["DeviceLocations"]:
            if fgt_dict['fgt_name'] == device['Upstream FortiGate Name'] and \
                    device['Device Type'] == 'FortiSwitch':
                fsw_dict = {'fortiswitch_name': device['Device Name'],
                            'fortiswitch_platform': device['Device Model'],
                            'fortiswitch_serial_number': device['Device Serial Number'],
                            'upstream_fortigate_name': device['Upstream FortiGate Name'],
                            'ports': []}
                fgt_dict['fortiswitches'].append(fsw_dict)

        # get the fortiswitch port data
        for fsw in fgt_dict['fortiswitches']:
            for fswport in sheets['FortiSwitchPorts']:
                if fswport['FortiSwitch Name'] == fsw['fortiswitch_name']:
                    fsw_port_dict = format_excel_columns(fswport)
                    if fsw_port_dict['allowed_vlans']:
                        if "," in fsw_port_dict['allowed_vlans']:
                            allowed_vlans = fsw_port_dict['allowed_vlans'].split(',')
                        elif "\n" in fsw_port_dict['allowed_vlans']:
                            allowed_vlans = fsw_port_dict['allowed_vlans'].split('\n')
                        else:
                            allowed_vlans = [fsw_port_dict['allowed_vlans']]
                        fsw_port_dict['allowed_vlans'] = allowed_vlans
                    fsw['ports'].append(fsw_port_dict)

        # get any APs downstream from this fortigate
        fgt_dict['fortiaps'] = []
        for device in sheets['DeviceLocations']:
            if fgt_dict['fgt_name'] == device['Upstream FortiGate Name'] and \
                    device['Device Type'] == 'FortiAP':
                # fap_dict = format_excel_columns(fap)
                fap_dict = {'fortiap_name': device['Device Name'],
                            'fortiap_platform': device['Device Model'],
                            'fortiap_serial_number': device['Device Serial Number'],
                            'fap_profile': None,
                            'upstream_fortigate_name': device['Upstream FortiGate Name']}
                for fap_profile in sheets["AP_Profiles"]:
                    # look for a matching FortiAP Profile by Location Name and AP Platform
                    if fgt_dict['fgt_location_name'] == fap_profile["Location Name"]:
                        if fap_dict['fortiap_platform'] == fap_profile['FortiAP Platform']:
                            fap_profile_dict = format_excel_columns(fap_profile)
                            profile_exists = False
                            for existing_ap_profile in fgt_dict['fap_profiles']:
                                if fap_profile_dict['profile_name'] == existing_ap_profile['profile_name']:
                                    profile_exists = True
                            if not profile_exists:
                                fgt_dict['fap_profiles'].append(fap_profile_dict)
                            fap_dict['fap_profile'] = fap_profile_dict['profile_name']
                fgt_dict['fortiaps'].append(fap_dict)

        for fap_profile in fgt_dict['fap_profiles']:
            if fap_profile['radio_1_channels']:
                if "," in fap_profile['radio_1_channels']:
                    r1_channels = fap_profile['radio_1_channels'].split(',')
                elif "\n" in fap_profile['radio_1_channels']:
                    r1_channels = fap_profile['radio_1_channels'].split('\n')
                else:
                    r1_channels = [fap_profile['radio_1_channels']]
                fap_profile['radio_1_channels'] = r1_channels

            if fap_profile['radio_2_channels']:
                if "," in fap_profile['radio_2_channels']:
                    r2_channels = fap_profile['radio_2_channels'].split(',')
                elif "\n" in fap_profile['radio_2_channels']:
                    r2_channels = fap_profile['radio_2_channels'].split('\n')
                else:
                    r2_channels = [fap_profile['radio_2_channels']]
                fap_profile['radio_2_channels'] = r2_channels

        # get the SSIDs assigned to the FortiGate LOCATION (we assume all FAPs in that location in all SSIDs
        # you can edit this later in the resulting files under {{role_path}}/vars/fortigates/
        fgt_dict["fap_ssids"] = []
        for ssid in sheets['SSIDs']:
            if fgt_dict['fgt_name'] == ssid['FortiGate Name']:
                if ssid['Captive Exempt CIDRs']:
                    if "," in ssid['Captive Exempt CIDRs']:
                        exempts = ssid['Captive Exempt CIDRs'].split(',')
                    elif '\n' in ssid['Captive Exempt CIDRs']:
                        exempts = ssid['Captive Exempt CIDRs'].split('\n')
                    else:
                        exempts = [ssid['Captive Exempt CIDRs']]
                    exempts = [x.strip() for x in exempts]
                    ssid['Captive Exempt CIDRs'] = exempts
                fgt_dict["fap_ssids"].append(format_excel_columns(ssid))

        fgt_dict['radius_servers'] = []
        for radius in sheets['Radius_Servers']:
            if fgt_dict['fgt_name'] == radius['FortiGate Name']:
                fgt_dict['radius_servers'].append(format_excel_columns(radius))

        # now get the ipsec vpns assigned to this fortigate
        fgt_dict["ipsec"] = []
        for ipsec in sheets['IPSec']:
            if fgt_dict['fgt_name'] == ipsec['FortiGate Name']:
                if ipsec['Allow Access']:
                    if "," in ipsec['Allow Access']:
                        allow_access = ipsec['Allow Access'].split(',')
                    elif "\n" in ipsec['Allow Access']:
                        allow_access = ipsec['Allow Access'].split('\n')
                    else:
                        allow_access = [ipsec['Allow Access']]
                    allow_access = [x.strip() for x in allow_access]
                    ipsec['Allow Access'] = allow_access
                fgt_dict['ipsec'].append(format_excel_columns(ipsec))

        # now get sdwan interfaces assigned to this fortigate
        fgt_dict['sdwan_interfaces'] = []
        for sdwan_interface in sheets['SDWAN_Interfaces']:
            if fgt_dict['fgt_name'] == sdwan_interface['FortiGate Name']:
                fgt_dict['sdwan_interfaces'].append(format_excel_columns(sdwan_interface))

        # now loop through any sdwan interfaces to determine the zones
        fgt_dict['sdwan_zones'] = []
        for sdwan_interface in fgt_dict['sdwan_interfaces']:
            if sdwan_interface['zone'] not in fgt_dict['sdwan_zones']:
                fgt_dict['sdwan_zones'].append(sdwan_interface['zone'])

        # now get the SDWAN Rules
        fgt_dict['sdwan_rules'] = []
        for sdwan_rule in sheets['SDWAN_Rules']:
            if fgt_dict['fgt_name'] == sdwan_rule['FortiGate Name']:
                fgt_dict['sdwan_rules'].append(format_excel_columns(sdwan_rule))

        # now get the SDWAN SLAs
        fgt_dict['sdwan_slas'] = []
        for sdwan_sla in sheets['SDWAN_SLAs']:
            if fgt_dict['fgt_name'] == sdwan_sla['FortiGate Name']:
                fgt_dict['sdwan_slas'].append(format_excel_columns(sdwan_sla))

        # BGP!
        # get the bgp neighbors
        fgt_dict['bgp_neighbors'] = []
        for neighbor in sheets['BGP_Neighbors']:
            if fgt_dict['fgt_name'] == neighbor['FortiGate Name']:
                fgt_dict['bgp_neighbors'].append(format_excel_columns(neighbor))
                if not fgt_dict['fgt_local_bgp_as']:
                    fgt_dict['fgt_local_bgp_as'] = neighbor['Local AS']

        # get any bgp neighbor groups
        fgt_dict['bgp_neighbor_groups'] = []
        for neighbor_group in sheets['BGP_Neighbor_Groups']:
            if fgt_dict['fgt_name'] == neighbor_group['FortiGate Name']:
                fgt_dict['bgp_neighbor_groups'].append(format_excel_columns(neighbor_group))

        # get bgp route maps
        fgt_dict['bgp_route_maps'] = []
        for route_map in sheets['BGP_Route_Maps']:
            if fgt_dict['fgt_name'] == route_map['FortiGate Name']:
                if route_map['Set Community']:
                    if "," in route_map['Set Community']:
                        set_community = route_map['Set Community'].split(',')
                    elif "\n" in route_map['Set Community']:
                        set_community = route_map['Set Community'].split("\n")
                    else:
                        set_community = [route_map['Set Community']]
                    set_community = [x.strip() for x in set_community]
                    route_map['Set Community'] = set_community
                fgt_dict['bgp_route_maps'].append(format_excel_columns(route_map))

        # get bgp community lists
        fgt_dict['bgp_community_lists'] = []
        for community_list in sheets['BGP_Community_Lists']:
            if fgt_dict['fgt_name'] == community_list['FortiGate Name']:
                fgt_dict['bgp_community_lists'].append(format_excel_columns(community_list))

        # get AS PATH LISTS
        fgt_dict['as_path_lists'] = []
        for aspath_list in sheets['ASPath_Lists']:
            if fgt_dict['fgt_name'] == aspath_list['FortiGate Name']:
                fgt_dict['as_path_lists'].append(format_excel_columns(aspath_list))

        # get route prefix lists
        fgt_dict['prefix_lists'] = []
        for prefix_list in sheets['Prefix_Lists']:
            if fgt_dict['fgt_name'] == prefix_list['FortiGate Name']:
                fgt_dict['prefix_lists'].append(format_excel_columns(prefix_list))

        # now get any policy packages
        fgt_dict['policy_packages'] = []
        for package in sheets['Policy_Packages']:
            if fgt_dict['fgt_name'] == package['FortiGate Name']:
                package['policies'] = []
                fgt_dict['policy_packages'].append(format_excel_columns(package))

        # now get the policies adom packages
        for policy in sheets['ADOM_Policies']:
            for package in fgt_dict['policy_packages']:
                if policy['Policy Package'] == package['policy_package']:
                    package['policies'].append(format_excel_columns(policy))

        # now get the HA settings
        fgt_dict['ha'] = []
        for ha in sheets['HA']:
            if fgt_dict['fgt_name'] == ha['FortiGate Name']:
                if ha['Monitor Ports']:
                    if not isinstance(ha['Monitor Ports'], list):
                        if "," in ha['Monitor Ports']:
                            monports = ha['Monitor Ports'].split(',')
                        elif '\n' in ha['Monitor Ports']:
                            monports = ha['Monitor Ports'].split('\n')
                        else:
                            monports = [ha['Monitor Ports']]
                        monports = [x.strip() for x in monports]
                        ha['Monitor Ports'] = monports
                else:
                    ha['Monitor Ports'] = []

                if not isinstance(ha['HA Ports'], list):
                    if "," in ha['HA Ports']:
                        haports = ha['HA Ports'].split(',')
                    elif '\n' in ha['HA Ports']:
                        haports = ha['HA Ports'].split('\n')
                    else:
                        haports = [ha['HA Ports']]
                    haports = [x.strip() for x in haports]
                    x = 0
                    haports_indexes = []
                    for port in haports:
                        haports_indexes.append(port + " " + str(x))
                        x += 1
                    ha['HA Ports'] = haports_indexes
                fgt_dict['ha'].append(format_excel_columns(ha))

        # now get the address groups
        fgt_dict['address_groups'] = []
        for addr_grp in sheets['Address_Groups']:
            addr_grp['adom'] = fgt_dict['fmgr_target_adom']
            if not isinstance(addr_grp['Subnets'], list):
                if "," in addr_grp['Subnets']:
                    subnets = addr_grp['Subnets'].split(',')
                elif "\n" in addr_grp['Subnets']:
                    subnets = addr_grp['Subnets'].split('\n')
                else:
                    subnets = [addr_grp['Subnets']]
                subnets = [x.strip() for x in subnets]
                addr_grp['Subnets'] = subnets
            fgt_dict['address_groups'].append(format_excel_columns(addr_grp))

        # now get the service groups
        fgt_dict['service_groups'] = []
        for svc_grp in sheets['Service_Groups']:
            svc_grp['adom'] = fgt_dict['fmgr_target_adom']
            if not isinstance(svc_grp['Ports'], list):
                if "," in svc_grp['Ports']:
                    ports = svc_grp['Ports'].split(',')
                elif "\n" in svc_grp['Ports']:
                    ports = svc_grp['Ports'].split('\n')
                else:
                    ports = [svc_grp['Ports']]
                ports = [x.strip() for x in ports]
                svc_grp['Ports'] = ports
            fgt_dict['service_groups'].append(format_excel_columns(svc_grp))

        ztp_fortigates.append(fgt_dict)
    return ztp_fortigates


def main():
    parser = argparse.ArgumentParser(description="Ansible ZTP Excel Workbook Importer")
    parser.add_argument("--file",
                        default="../spreadsheets/Ansible_ZTP_Local_Simple_Topology.xlsx",
                        # default="../spreadsheets/Evoke_SDWAN_Demo.xlsx",
                        # default="../spreadsheets/Accenture_ZTP.xlsx",
                        help="File path for Excel XLSX Workbook")
    parser.add_argument("--role_vars_dir",
                        default=None,
                        help="Directory Path to the Ansible ZTP 'Vars' folder")
    args = parser.parse_args()
    print("Using workbook: " + args.file)

    sheets = convert_whole_workbook_to_dict(workbook_path=args.file)
    ztp_fortigates = generate_fortigate_var_files(sheets=sheets)

    input_file_name = os.path.basename(args.file)
    dict_key_name = input_file_name.replace('.xlsx', '')
    whole_worksheet_out_dict = {dict_key_name: sheets}

    if args.role_vars_dir:
        # make the "main" and "fortigates" sub directories under {{role_path}}/vars/
        try:
            os.mkdir(args.role_vars_dir + "/main")
        except FileExistsError:
            # if the out directory already exists, then move on.
            pass
        try:
            os.mkdir(args.role_vars_dir + "/fortigates")
        except FileExistsError:
            # if the out directory already exists, then move on.
            pass
        # write the entire spreadsheet import to a file
        # with the name of the spreadsheet as the variable name
        # allows you to access the entire spreadsheet from ansible, not just what we collected with
        # the generate_fortigate_var_files() function
        # with open(args.role_vars_dir + "/main/" + "imported_" + input_file_name.replace(".xlsx", ".yaml"), 'w') as outfile:
        #     yaml.dump(whole_worksheet_out_dict, outfile)

        # now we need to write every fortigate we found under ztp_fortigates to their own yaml files
        for fortigate in ztp_fortigates:
            out_file_name = fortigate['fgt_name'] + ".yaml"
            with open(args.role_vars_dir + "/fortigates/" + out_file_name, 'w') as outfortigate:
                yaml.dump(fortigate, outfortigate)

    # if no output directory was specified, write the spreadsheet yaml to screen.
    elif not args.role_vars_dir:
        print(yaml.dump(whole_worksheet_out_dict))
        print("We found " + str(len(ztp_fortigates)) + " FortiGate(s) within this dataset.")


if __name__ == '__main__':
    main()
